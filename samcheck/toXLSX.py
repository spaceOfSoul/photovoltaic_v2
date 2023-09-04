import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

df = pd.read_csv('pv_f16.csv')  

df['일시'] = pd.to_datetime(df['일시'])

df.set_index('date', inplace=True)

ac_tot = df['누적발전량']

ac_tot_hourly = ac_tot.resample('H').last() - ac_tot.resample('H').first()

all_dates = pd.date_range(start='2022-01-01 00:00:00', end='2022-08-31 23:00:00', freq='H')
missing_df = pd.DataFrame(index=all_dates)
df_hourly = pd.merge(missing_df, ac_tot_hourly, how='left', left_index=True, right_index=True).fillna(0)

groups = df_hourly.groupby(df_hourly.index.date)

for name, group in groups:
    year = name.year
    month = name.month
    day = name.day

    if year == 2021:
        continue

    directory = os.path.join('data/', f'{month}월')
    if not os.path.exists(directory):
        os.makedirs(directory)

    file_name = f'{day}.xlsx'

    #Workbook : excel을 직접 건드리는데에 필요.
    wb = Workbook()
    ws = wb.active

    ws.append(['시간 발전량'])
    ws.append([])
    ws.append([f'[ {year}년 {month}월 {day}일 ]'])

    for r in dataframe_to_rows(group, index=True, header=False):
        ws.append(r)

    total_row = ws.max_row + 1
    ws[f'A{total_row}'] = '총합'
    ws[f'B{total_row}'] = f'=SUM(B4:B{total_row-1})'

    wb.save(os.path.join(directory, file_name))
